from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import os
import io
import csv
from pathlib import Path
from urllib.parse import quote

app = Flask(__name__)
app.secret_key = 'eduflow-secret-key-2026'

# Configuration de la base de données SQLite
basedir = os.path.abspath(os.path.dirname(__file__))
db_file = 'etudiants.db'
db_path = os.path.join(basedir, db_file)

# Configuration avec le chemin URLencoded
db_uri = 'sqlite:///' + db_path.replace('\\', '/').replace(' ', '%20')

app.config['SQLALCHEMY_DATABASE_URI'] = db_uri
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'connect_args': {'check_same_thread': False, 'timeout': 30},
    'pool_pre_ping': True,
    'pool_recycle': 3600,
}

# Initialiser la base de données
db = SQLAlchemy(app)


# Modèle Étudiant
class Etudiant(db.Model):
    __tablename__ = 'etudiants'
    
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    heure = db.Column(db.String(20), nullable=False)
    date = db.Column(db.Date, nullable=False, default=datetime.today)
    statut = db.Column(db.String(10), nullable=False, default='Présent')  # Présent ou Absent
    
    def __repr__(self):
        return f'<Étudiant {self.nom} - {self.statut}>'


# Créer les tables au démarrage
def init_db():
    """Initialiser la base de données"""
    try:
        with app.app_context():
            db.create_all()
            print(f"[OK] Base de donnees initialisee : {db_path}")
    except Exception as e:
        print(f"[ERR] Erreur lors de l'initialisation : {e}")


def save_to_db(etudiant):
    """
    Fonction sécurisée pour sauvegarder les données
    """
    try:
        db.session.add(etudiant)
        db.session.flush()  # Vérifier la validité avant le commit
        db.session.commit()
        return True
    except Exception as e:
        db.session.rollback()
        print(f"[ERR] Erreur lors de la sauvegarde : {e}")
        return False


# Initialiser la BD au démarrage
init_db()


@app.route('/')
def accueil():
    """
    Route d'accueil : affiche la liste complète des étudiants présents
    """
    presences = Etudiant.query.order_by(Etudiant.id.desc()).all()
    total = len(presences)
    return render_template('index.html', presences=presences, total=total)


@app.route('/pointage/<nom>')
def pointage_rapide(nom):
    """
    Route dynamique : إضافة مباشرة إلى البيانات - من المفروض أنه حاضر
    """
    if nom.strip():
        try:
            nouvel_etudiant = Etudiant(
                nom=nom.strip(),
                heure=datetime.now().strftime('%H:%M:%S'),
                date=datetime.now().date(),
                statut='Présent'
            )
            if save_to_db(nouvel_etudiant):
                print(f"[OK] {nom} ajoute avec succes - Present")
            else:
                print(f"[ERR] Erreur lors de l'ajout de {nom}")
        except Exception as e:
            print(f"[ERR] Exception: {e}")
    return redirect(url_for('accueil'))


@app.route('/ajouter', methods=['GET', 'POST'])
def ajouter():
    """
    الإضافة عبر النموذج
    """
    if request.method == 'POST':
        nom = request.form.get('nom', '').strip()
        statut = request.form.get('statut', 'Présent').strip()
        
        if nom:
            try:
                nouvel_etudiant = Etudiant(
                    nom=nom,
                    heure=datetime.now().strftime('%H:%M:%S'),
                    date=datetime.now().date(),
                    statut=statut if statut in ['Présent', 'Absent'] else 'Présent'
                )
                if save_to_db(nouvel_etudiant):
                    print(f"[OK] {nom} enregistre avec succes - Statut: {statut}")
                    return redirect(url_for('accueil'))
                else:
                    flash("Erreur lors de l'enregistrement!", 'error')
                    return render_template('formulaire.html', erreur="Erreur de base de données")
            except Exception as e:
                print(f"[ERR] Exception: {e}")
                flash(f"Erreur: {str(e)}", 'error')
                return render_template('formulaire.html', erreur=str(e))
        else:
            return render_template('formulaire.html', erreur="Le nom ne peut pas être vide!")
    
    return render_template('formulaire.html')


@app.route('/reinitialiser')
def reinitialiser():
    """
    حذف جميع البيانات من قاعدة البيانات
    """
    try:
        db.session.query(Etudiant).delete()
        db.session.commit()
        print("[OK] Base de donnees nettoyee")
    except Exception as e:
        db.session.rollback()
        print(f"[ERR] Erreur lors du nettoyage: {e}")
    return redirect(url_for('accueil'))


@app.route('/modifier/<int:id>', methods=['GET', 'POST'])
def modifier(id):
    """
    تعديل بيانات طالب
    """
    etudiant = Etudiant.query.get_or_404(id)
    
    if request.method == 'POST':
        nom = request.form.get('nom', '').strip()
        statut = request.form.get('statut', '').strip()
        
        if nom:
            try:
                etudiant.nom = nom
                if statut in ['Présent', 'Absent']:
                    etudiant.statut = statut
                db.session.commit()
                print(f"[OK] {nom} modifie avec succes")
                flash(f"Étudiant {nom} modifié avec succès!", 'success')
                return redirect(request.form.get('redirect_to', url_for('accueil')))
            except Exception as e:
                db.session.rollback()
                print(f"[ERR] Erreur lors de la modification: {e}")
                flash(f"Erreur: {str(e)}", 'error')
        else:
            flash("Le nom ne peut pas être vide!", 'error')
    
    redirect_to = request.args.get('redirect_to', request.referrer or url_for('accueil'))
    return render_template('modifier.html', etudiant=etudiant, redirect_to=redirect_to)


@app.route('/supprimer/<int:id>')
def supprimer(id):
    """
    حذف طالب من الجدول
    """
    redirect_to = request.args.get('redirect_to', request.referrer or url_for('accueil'))
    try:
        etudiant = Etudiant.query.get_or_404(id)
        nom = etudiant.nom
        db.session.delete(etudiant)
        db.session.commit()
        print(f"[OK] {nom} supprime")
        flash(f"Étudiant {nom} supprimé avec succès!", 'success')
    except Exception as e:
        db.session.rollback()
        print(f"[ERR] Erreur lors de la suppression: {e}")
        flash(f"Erreur lors de la suppression: {str(e)}", 'error')
    return redirect(redirect_to)


@app.errorhandler(404)
def page_not_found(error):
    """
    Gestion personnalisée de l'erreur 404
    """
    return render_template('erreur_404.html'), 404


@app.route('/statistiques')
def statistiques():
    """
    Affiche les statistiques des presences avec graphiques et stats avancées
    """
    etudiants = Etudiant.query.all()
    total_etudiants = len(etudiants)
    
    # Compter par étudiant (présences et absences séparément)
    etudiants_presences = {}
    etudiants_absences = {}
    etudiants_total = {}
    presences_count = 0
    absences_count = 0
    
    for etudiant in etudiants:
        nom = etudiant.nom
        if nom not in etudiants_presences:
            etudiants_presences[nom] = 0
            etudiants_absences[nom] = 0
            etudiants_total[nom] = 0
        
        etudiants_total[nom] += 1
        
        if etudiant.statut == 'Présent':
            etudiants_presences[nom] += 1
            presences_count += 1
        else:
            etudiants_absences[nom] += 1
            absences_count += 1
    
    # Calculer le pourcentage de présence pour chaque étudiant
    etudiants_pourcentage = {}
    for nom in etudiants_total:
        total = etudiants_total[nom]
        present = etudiants_presences.get(nom, 0)
        etudiants_pourcentage[nom] = round((present / total) * 100, 1) if total > 0 else 0
    
    # Trier par pourcentage de présence (décroissant)
    etudiants_pourcentage_sorted = dict(
        sorted(etudiants_pourcentage.items(), key=lambda x: x[1], reverse=True)
    )
    
    # Statistiques avancées : plus présent et plus absent
    plus_present = None
    plus_absent = None
    
    if etudiants_presences:
        plus_present_nom = max(etudiants_presences, key=etudiants_presences.get)
        plus_present = {
            'nom': plus_present_nom,
            'count': etudiants_presences[plus_present_nom],
            'pourcentage': etudiants_pourcentage.get(plus_present_nom, 0)
        }
    
    if etudiants_absences:
        plus_absent_nom = max(etudiants_absences, key=etudiants_absences.get)
        plus_absent = {
            'nom': plus_absent_nom,
            'count': etudiants_absences[plus_absent_nom],
            'pourcentage': round(100 - etudiants_pourcentage.get(plus_absent_nom, 100), 1)
        }
    
    # Données pour Chart.js
    chart_labels = list(etudiants_pourcentage_sorted.keys())
    chart_presences = [etudiants_presences.get(nom, 0) for nom in chart_labels]
    chart_absences = [etudiants_absences.get(nom, 0) for nom in chart_labels]
    chart_pourcentages = [etudiants_pourcentage_sorted[nom] for nom in chart_labels]
    
    return render_template('statistiques.html', 
                         total=total_etudiants, 
                         etudiants_counts=etudiants_total,
                         etudiants_pourcentage=etudiants_pourcentage_sorted,
                         presences_count=presences_count,
                         absences_count=absences_count,
                         plus_present=plus_present,
                         plus_absent=plus_absent,
                         chart_labels=chart_labels,
                         chart_presences=chart_presences,
                         chart_absences=chart_absences,
                         chart_pourcentages=chart_pourcentages)


@app.route('/historique')
def historique():
    """
    Affiche l'historique avec recherche et filtrage
    """
    # Récupérer les paramètres de recherche/filtrage
    search = request.args.get('search', '').strip()
    statut_filter = request.args.get('statut', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    
    # Construire la requête
    query = Etudiant.query
    
    if search:
        query = query.filter(Etudiant.nom.ilike(f'%{search}%'))
    
    if statut_filter and statut_filter in ['Présent', 'Absent']:
        query = query.filter(Etudiant.statut == statut_filter)
    
    if date_from:
        try:
            date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Etudiant.date >= date_from_parsed)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Etudiant.date <= date_to_parsed)
        except ValueError:
            pass
    
    etudiants = query.order_by(Etudiant.date.desc(), Etudiant.id.desc()).all()
    
    return render_template('historique.html', 
                         presences=etudiants,
                         search=search,
                         statut_filter=statut_filter,
                         date_from=date_from,
                         date_to=date_to)


# ========================
# Routes d'exportation
# ========================

@app.route('/export/csv')
def export_csv():
    """Exporter les données en CSV"""
    etudiants = Etudiant.query.order_by(Etudiant.date.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # En-tête
    writer.writerow(['ID', 'Nom', 'Statut', 'Date', 'Heure'])
    
    # Données
    for e in etudiants:
        writer.writerow([
            e.id,
            e.nom,
            e.statut,
            e.date.strftime('%d/%m/%Y') if e.date else '',
            e.heure
        ])
    
    output.seek(0)
    
    # Ajouter BOM UTF-8 pour Excel
    bom = '\ufeff'
    csv_content = bom + output.getvalue()
    
    return Response(
        csv_content,
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f'attachment; filename=presences_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        }
    )


@app.route('/export/excel')
def export_excel():
    """Exporter les données en Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Présences"
    
    # Styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titre
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f"Rapport de Présences - EduFlow-Check ({datetime.now().strftime('%d/%m/%Y')})"
    title_cell.font = Font(name='Arial', size=14, bold=True, color='667EEA')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    # En-têtes
    headers = ['ID', 'Nom', 'Statut', 'Date', 'Heure']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Données
    etudiants = Etudiant.query.order_by(Etudiant.date.desc()).all()
    
    present_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    absent_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    
    for row, e in enumerate(etudiants, 4):
        fill = present_fill if e.statut == 'Présent' else absent_fill
        
        data = [e.id, e.nom, e.statut, e.date.strftime('%d/%m/%Y') if e.date else '', e.heure]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    
    # Sauvegarder dans un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'presences_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@app.route('/export/pdf')
def export_pdf():
    """Exporter les données en PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=1,
        textColor=colors.HexColor('#667EEA')
    )
    elements.append(Paragraph("Rapport de Presences - EduFlow-Check", title_style))
    elements.append(Paragraph(f"Date: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Données du tableau
    etudiants = Etudiant.query.order_by(Etudiant.date.desc()).all()
    
    table_data = [['#', 'Nom', 'Statut', 'Date', 'Heure']]
    for i, e in enumerate(etudiants, 1):
        table_data.append([
            str(i),
            e.nom,
            e.statut,
            e.date.strftime('%d/%m/%Y') if e.date else '',
            e.heure
        ])
    
    # Créer le tableau
    col_widths = [1.5*cm, 8*cm, 4*cm, 4*cm, 4*cm]
    table = Table(table_data, colWidths=col_widths)
    
    # Style du tableau
    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
    ]
    
    # Couleurs alternées + couleur selon statut
    for i, e in enumerate(etudiants, 1):
        if e.statut == 'Absent':
            style_commands.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFF5F5')))
        elif i % 2 == 0:
            style_commands.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F7FAFC')))
    
    table.setStyle(TableStyle(style_commands))
    elements.append(table)
    
    # Résumé en bas
    elements.append(Spacer(1, 30))
    total_presences = sum(1 for e in etudiants if e.statut == 'Présent')
    total_absences = sum(1 for e in etudiants if e.statut != 'Présent')
    
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=5,
        textColor=colors.HexColor('#4A5568')
    )
    elements.append(Paragraph(f"Total: {len(etudiants)} | Presences: {total_presences} | Absences: {total_absences}", summary_style))
    
    doc.build(elements)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'rapport_presences_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )


# ========================
# Routes API existantes
# ========================

@app.route('/api/etudiants')
def api_etudiants():
    """API JSON pour obtenir tous الطلاب"""
    etudiants = Etudiant.query.all()
    return jsonify([{
        'id': e.id,
        'nom': e.nom,
        'heure': e.heure,
        'date': str(e.date),
        'statut': e.statut
    } for e in etudiants])


@app.route('/api/export')
def api_export():
    """Exporte les données en JSON"""
    etudiants = Etudiant.query.all()
    data = {
        'total': len(etudiants),
        'export_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'etudiants': [{
            'id': e.id,
            'nom': e.nom,
            'heure': e.heure,
            'date': str(e.date),
            'statut': e.statut
        } for e in etudiants]
    }
    return jsonify(data)


@app.route('/api/statut')
def api_statut():
    """حالة قاعدة البيانات والتشخيص"""
    db_exists = os.path.exists(db_path)
    db_size = os.path.getsize(db_path) if db_exists else 0
    
    try:
        total = Etudiant.query.count()
        status = 'active'
    except Exception as e:
        total = -1
        status = f'error: {str(e)}'
    
    return jsonify({
        'status': status,
        'database_file': db_path,
        'database_exists': db_exists,
        'database_size_bytes': db_size,
        'total_etudiants': total,
        'last_update': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'uri': app.config['SQLALCHEMY_DATABASE_URI']
    })


@app.route('/api/supprimer/<int:id>', methods=['POST'])
def api_supprimer(id):
    """حذف طالب محدد"""
    try:
        etudiant = Etudiant.query.get(id)
        if etudiant:
            db.session.delete(etudiant)
            db.session.commit()
            print(f"[OK] {etudiant.nom} supprime")
            return jsonify({'success': True, 'message': f'Étudiant {etudiant.nom} supprimé'})
        return jsonify({'success': False, 'message': 'Étudiant non trouvé'}), 404
    except Exception as e:
        db.session.rollback()
        print(f"[ERR] Erreur lors de la suppression: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, host='localhost', port=5000)
